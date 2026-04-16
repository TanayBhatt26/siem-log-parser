"""exporters/csv_exporter.py — CSV and Excel (.xlsx) exporters"""

import csv, io
from typing import List
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
from schema import LogEvent

COLUMNS = [
    "event_id","timestamp","source_format","source_host","source_ip","source_port",
    "dest_ip","dest_port","event_type","event_action","severity","severity_code",
    "category","username","user_id","process_name","process_id","protocol",
    "bytes_in","bytes_out","message",
]

def to_csv(events: List[LogEvent]) -> str:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for evt in events:
        writer.writerow({k: getattr(evt, k, "") for k in COLUMNS})
    return buf.getvalue()

def to_excel(events: List[LogEvent]) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIEM Events"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col.replace("_"," ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    sev_colors = {"critical":"FF0000","high":"FF6600","medium":"FFCC00","low":"00CC44"}

    for row_idx, evt in enumerate(events, 2):
        row_data = {k: getattr(evt, k, "") for k in COLUMNS}
        sev = str(row_data.get("severity","")).lower()
        for col_idx, col in enumerate(COLUMNS, 1):
            val = row_data[col]
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            if sev_colors.get(sev) and col == "severity":
                cell.fill = PatternFill("solid", fgColor=sev_colors[sev])
                cell.font = Font(bold=True, color="FFFFFF")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
